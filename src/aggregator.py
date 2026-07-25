"""Cliente do agregador bancário — abstração única com 3 modos:

  "pluggy"/"belvo": Open Finance via API paga (ver README, seção "Agregador
                    bancário" — planos comerciais custam R$1.500-6.000+/mês,
                    não vale a pena pro porte deste negócio).
  "arquivo": modo grátis — lê extratos exportados manualmente do internet
             banking de cada conta (OFX ou XLSX, ver EXTRATOS_DIR). Exige que
             alguém baixe e salve o arquivo todo dia antes do Cenário 1 rodar
             (06:00), mas não tem custo nenhum.

Cobre as 6 contas reais (config.CONTAS_BANCARIAS). Retorna transações
normalizadas com a CHAVE da conta (ex: "AZEVEDO_ITAU"), não apenas "CONTA1/2/3".

  {id, conta_key, unidade, data, valor, tipo(CREDITO|DEBITO), descricao,
   pix_key_destino}

IMPORTANTE (confirmado com extrato real do Sicoob [XLSX e PDF] e do Itaú
[XLSX]): o extrato bancário não expõe em nenhum campo "qual chave Pix foi
usada para receber" — o Sicoob mascara o CPF/CNPJ de quem paga e de quem
recebe; o Itaú mostra Nome/Razão Social e CPF/CNPJ completos, mas também sem
a chave Pix usada. Ou seja, `pix_key_destino` NUNCA vem preenchido nesses
formatos — a identificação de empresa na conta Azevedo (compartilhada Casa
da Árvore + Casarão) depende inteiramente do matching por valor/vencimento
contra Contas_a_Receber (protegido pela trava de ambiguidade do Cenário 1).

Sicoob e Itaú exportam XLSX em layouts BEM diferentes (ver
`_xlsx_transacoes_sicoob` vs `_xlsx_transacoes_itau`) — o dispatcher
`_xlsx_transacoes` escolhe o parser certo por `config.CONTAS_BANCARIAS[conta_key]["banco"]`.
A Caixa exporta OFX nativo (mais confiável que os PDFs/OCR que recebemos
antes) — mas reaproveita o mesmo FITID pra várias transações do mesmo lote,
por isso `_ofx_transacoes` deduplica por FITID+data+valor+memo, não só FITID.
"""
import hashlib, json, os, re, sys, requests
from datetime import date, datetime, timedelta
from pathlib import Path
from dotenv import load_dotenv
import config
from .net_utils import HTTP_TIMEOUT, com_retry
from .logging_config import get_logger

logger = get_logger(__name__)

load_dotenv()
AGG = os.getenv("AGGREGATOR", "pluggy")

# resolve link_id de cada conta a partir do .env (modos pluggy/belvo)
LINKS = {key: os.getenv(env_var)
         for key, env_var in config.LINK_ID_ENV_VAR.items()}

# modo "arquivo": pasta com um .ofx por conta (nome = conta_key, ex:
# extratos/AZEVEDO_SICOOB.ofx) + um registro local dos FITIDs já processados,
# pra não contar a mesma transação duas vezes quando o extrato exportado
# de novo ainda inclui dias já processados.
EXTRATOS_DIR = Path(os.getenv("EXTRATOS_DIR", "extratos"))
PROCESSADOS_LOG = EXTRATOS_DIR / ".processados.json"


@com_retry()
def _pluggy_token():
    r = requests.post("https://api.pluggy.ai/auth", json={
        "clientId": os.getenv("PLUGGY_CLIENT_ID"),
        "clientSecret": os.getenv("PLUGGY_CLIENT_SECRET")}, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    return r.json()["apiKey"]


@com_retry()
def _pluggy_transacoes(link_id, desde):
    tk = _pluggy_token()
    out, page = [], 1
    while True:
        r = requests.get("https://api.pluggy.ai/transactions",
                         headers={"X-API-KEY": tk},
                         params={"accountId": link_id, "from": desde.isoformat(),
                                 "pageSize": 100, "page": page}, timeout=HTTP_TIMEOUT)
        r.raise_for_status()
        data = r.json()
        out += data.get("results", [])
        if page >= data.get("totalPages", 1):
            break
        page += 1
    return out


@com_retry()
def _belvo_transacoes(link_id, desde):
    auth = (os.getenv("BELVO_SECRET_ID"), os.getenv("BELVO_SECRET_PASSWORD"))
    r = requests.post("https://api.belvo.com/api/transactions/",
                      auth=auth, json={"link": link_id,
                                       "date_from": desde.isoformat(),
                                       "date_to": date.today().isoformat()}, timeout=HTTP_TIMEOUT)
    r.raise_for_status()
    return r.json()


def _carregar_processados():
    if PROCESSADOS_LOG.exists():
        with open(PROCESSADOS_LOG, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def _salvar_processados(processados):
    EXTRATOS_DIR.mkdir(exist_ok=True)
    with open(PROCESSADOS_LOG, "w", encoding="utf-8") as f:
        json.dump(processados, f)


def _pix_key_no_texto(texto):
    """Extrai a chave Pix da descrição/memo da transação, se alguma das
    chaves configuradas aparecer no texto (extratos OFX costumam trazer a
    chave Pix embutida no campo de descrição, não em campo estruturado)."""
    for emp in config.EMPRESAS.values():
        for chave in emp["pix_keys"]:
            if chave in texto:
                return chave
    return ""


def _normalizar_ofx(t, conta_key):
    unidade = config.CONTAS_BANCARIAS[conta_key]["unidade"]
    valor = float(t.amount)
    descricao = (getattr(t, "memo", None) or getattr(t, "payee", None) or "").strip()
    return {
        "id": t.id, "conta_key": conta_key, "unidade": unidade,
        "data": t.date.date().isoformat(), "valor": abs(valor),
        "tipo": "CREDITO" if valor > 0 else "DEBITO",
        "descricao": descricao,
        "pix_key_destino": _pix_key_no_texto(descricao),
    }


def _ofx_transacoes(conta_key, desde, processados):
    """Lê extratos/{conta_key}.ofx e deduplica contra o que já foi processado
    em execuções anteriores.

    Não usa só o FITID: confirmado com extrato real da Caixa que ela
    reaproveita o mesmo FITID pra várias transações diferentes do mesmo lote
    (157 transações reais, só 97 FITIDs distintos) — usar só o FITID faria o
    sistema descartar transações de verdade como se já tivessem sido
    processadas. A chave de dedup é FITID + data + valor + memo (único nos
    157 casos reais testados)."""
    caminho = EXTRATOS_DIR / f"{conta_key}.ofx"
    from ofxparse import OfxParser
    with open(caminho, "rb") as f:
        ofx = OfxParser.parse(f)
    vistos = set(processados.get(conta_key, []))
    novas = []
    for conta_ofx in ofx.accounts:
        for t in conta_ofx.statement.transactions:
            if t.date.date() < desde:
                continue
            memo = (getattr(t, "memo", None) or getattr(t, "payee", None) or "").strip()
            chave = f"{t.id}|{t.date.date().isoformat()}|{t.amount}|{memo}"
            if chave in vistos:
                continue
            vistos.add(chave)
            novas.append(_normalizar_ofx(t, conta_key))
    processados[conta_key] = list(vistos)
    return novas


def _linha_e_data(valor):
    if isinstance(valor, (datetime, date)):
        return True
    return bool(re.match(r"^\d{2}/\d{2}/\d{4}$", str(valor).strip()))


def _linha_para_data_iso(valor):
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    return datetime.strptime(str(valor).strip(), "%d/%m/%Y").date().isoformat()


def _valor_e_tipo_xlsx(bruto):
    """Extratos exportados costumam vir como '6.000,00 C' ou '-\\xa050,00 D'
    (às vezes com espaço entre o sinal e o número)."""
    texto = str(bruto).replace("\xa0", " ").strip()
    tipo = "CREDITO" if texto.upper().endswith("C") else "DEBITO"
    numero = texto[:-1].replace(" ", "").replace(".", "").replace(",", ".")
    return abs(float(numero)), tipo


def _xlsx_transacoes_sicoob(conta_key, desde, processados):
    """Lê extratos/{conta_key}.xlsx — formato observado no Sicoob: cada
    transação ocupa um bloco de linhas (1ª linha: data + histórico + valor;
    linhas seguintes sem data: detalhes — tipo Recebimento/Pagamento Pix,
    contraparte, memo). Linhas "SALDO DO DIA"/"SALDO ANTERIOR" são só
    referência de saldo, não movimentação, e são ignoradas.

    Esse formato não tem um ID único de transação (diferente do FITID do
    OFX), então o dedup usa um hash de data+valor+tipo+detalhes — único na
    prática, mas duas transações genuinamente idênticas (mesmo valor, mesmo
    dia, mesmo texto) colidiriam e a segunda seria perdida."""
    from openpyxl import load_workbook
    caminho = EXTRATOS_DIR / f"{conta_key}.xlsx"
    linhas = list(load_workbook(caminho, data_only=True).active.iter_rows(values_only=True))
    vistos = set(processados.get(conta_key, []))
    novas = []
    i = 0
    while i < len(linhas):
        row = list(linhas[i]) + [None] * max(0, 4 - len(linhas[i]))
        data_raw, _doc, historico, valor_raw = row[:4]
        if not (data_raw and historico and _linha_e_data(data_raw)
                and "SALDO" not in str(historico).upper()):
            i += 1
            continue
        detalhes = []
        j = i + 1
        while j < len(linhas) and not linhas[j][0]:
            for cel in linhas[j][1:3]:
                if cel:
                    detalhes.append(str(cel).strip())
            j += 1
        data_iso = _linha_para_data_iso(data_raw)
        if data_iso >= desde.isoformat():
            valor, tipo = _valor_e_tipo_xlsx(valor_raw)
            descricao = " | ".join(detalhes)
            tx_id = hashlib.sha1(
                f"{data_iso}|{valor}|{tipo}|{descricao}".encode("utf-8")).hexdigest()
            if tx_id not in vistos:
                vistos.add(tx_id)
                novas.append({
                    "id": tx_id, "conta_key": conta_key,
                    "unidade": config.CONTAS_BANCARIAS[conta_key]["unidade"],
                    "data": data_iso, "valor": valor, "tipo": tipo,
                    "descricao": descricao,
                    "pix_key_destino": _pix_key_no_texto(descricao),
                })
        i = j
    processados[conta_key] = list(vistos)
    return novas


def _valor_e_tipo_xlsx_itau(bruto):
    """No extrato do Itaú o valor já vem como número assinado (positivo =
    crédito, negativo = débito), não como texto com sufixo C/D — mas trata
    também o caso de vir como texto, que alguns exports fazem."""
    if isinstance(bruto, (int, float)):
        valor = float(bruto)
    else:
        texto = str(bruto).replace("\xa0", " ").strip()
        texto = texto.replace(".", "").replace(",", ".")
        valor = float(texto)
    return abs(valor), ("CREDITO" if valor > 0 else "DEBITO")


def _xlsx_transacoes_itau(conta_key, desde, processados):
    """Lê extratos/{conta_key}.xlsx — formato "Extrato de Lançamentos" do
    Itaú (confirmado com extrato real, agência 8595/conta 0043484-9):
    linhas de metadado no topo, depois uma linha de cabeçalho literal
    "Data | Lançamento | Razão Social | CPF/CNPJ | Valor (R$) | Saldo (R$)",
    e a partir dali UMA linha por transação (diferente do Sicoob, que usa
    blocos de várias linhas). Linhas "SALDO ANTERIOR"/"SALDO TOTAL
    DISPONÍVEL DIA" não têm valor de movimentação (só saldo) e são puladas.

    Sem FITID nesse formato — dedup por hash de
    data+lançamento+razão social+CPF/CNPJ+valor (mesma limitação já
    documentada no Sicoob: duas transações genuinamente idênticas no mesmo
    dia colidiriam e a segunda seria perdida)."""
    from openpyxl import load_workbook
    caminho = EXTRATOS_DIR / f"{conta_key}.xlsx"
    linhas = list(load_workbook(caminho, data_only=True).active.iter_rows(values_only=True))

    inicio = next((i for i, r in enumerate(linhas) if r and r[0] == "Data"), None)
    if inicio is None:
        logger.warning("Cabeçalho 'Data' não encontrado em %s — extrato do Itaú "
                        "pode ter mudado de layout", caminho)
        return []

    vistos = set(processados.get(conta_key, []))
    novas = []
    for row in linhas[inicio + 1:]:
        row = list(row) + [None] * max(0, 6 - len(row))
        data_raw, lancamento, razao_social, cpf_cnpj, valor_raw, _saldo = row[:6]
        if not data_raw or not _linha_e_data(data_raw) or valor_raw is None:
            continue  # cobre SALDO ANTERIOR / SALDO TOTAL DISPONÍVEL DIA (sem valor)
        data_iso = _linha_para_data_iso(data_raw)
        if data_iso < desde.isoformat():
            continue
        valor, tipo = _valor_e_tipo_xlsx_itau(valor_raw)
        descricao = " | ".join(str(p).strip() for p in (lancamento, razao_social, cpf_cnpj) if p)
        tx_id = hashlib.sha1(
            f"{data_iso}|{lancamento}|{razao_social}|{cpf_cnpj}|{valor}".encode("utf-8")).hexdigest()
        if tx_id not in vistos:
            vistos.add(tx_id)
            novas.append({
                "id": tx_id, "conta_key": conta_key,
                "unidade": config.CONTAS_BANCARIAS[conta_key]["unidade"],
                "data": data_iso, "valor": valor, "tipo": tipo,
                "descricao": descricao,
                "pix_key_destino": _pix_key_no_texto(descricao),
            })
    processados[conta_key] = list(vistos)
    return novas


def _xlsx_transacoes(conta_key, desde, processados):
    """Despacha para o parser certo conforme o banco da conta — Sicoob e
    Itaú exportam XLSX em layouts incompatíveis entre si (ver docstrings de
    cada parser). Bancos ainda não confirmados com extrato real (hoje: Caixa
    — o arquivo que recebemos veio de OCR corrompido, sem confiabilidade pra
    parsear) caem no parser do Sicoob como melhor palpite, com aviso no log."""
    banco = config.CONTAS_BANCARIAS[conta_key]["banco"]
    if banco == "Itaú":
        return _xlsx_transacoes_itau(conta_key, desde, processados)
    if banco != "Sicoob":
        logger.warning("Formato XLSX do banco %s (%s) ainda não confirmado com "
                        "extrato real — tentando como Sicoob", banco, conta_key)
    return _xlsx_transacoes_sicoob(conta_key, desde, processados)


def _normalizar(raw, conta_key):
    unidade = config.CONTAS_BANCARIAS[conta_key]["unidade"]
    if AGG == "pluggy":
        return {
            "id": raw["id"], "conta_key": conta_key, "unidade": unidade,
            "data": raw["date"][:10], "valor": abs(raw["amount"]),
            "tipo": "CREDITO" if raw["amount"] > 0 else "DEBITO",
            "descricao": raw.get("description", ""),
            # TODO: confirmar campo real da chave Pix recebedora no payload Pluggy
            "pix_key_destino": (raw.get("paymentData") or {}).get(
                "receiver", {}).get("pixKey", ""),
        }
    return {
        "id": raw["id"], "conta_key": conta_key, "unidade": unidade,
        "data": raw["value_date"] or raw["accounting_date"],
        "valor": abs(raw["amount"]),
        "tipo": "CREDITO" if raw["type"] == "INFLOW" else "DEBITO",
        "descricao": raw.get("description", ""),
        "pix_key_destino": "",  # TODO: mapear campo Belvo equivalente
    }


def transacoes_ultimas_24h():
    desde = date.today() - timedelta(days=1)
    todas = []
    if AGG == "arquivo":
        processados = _carregar_processados()
        for conta_key in config.CONTAS_BANCARIAS:
            if (EXTRATOS_DIR / f"{conta_key}.ofx").exists():
                todas += _ofx_transacoes(conta_key, desde, processados)
            elif (EXTRATOS_DIR / f"{conta_key}.xlsx").exists():
                todas += _xlsx_transacoes(conta_key, desde, processados)
            else:
                logger.warning("Sem arquivo .ofx/.xlsx para %s (%s) — pulando",
                                conta_key, EXTRATOS_DIR)
        _salvar_processados(processados)
        return todas
    fetch = _pluggy_transacoes if AGG == "pluggy" else _belvo_transacoes
    for conta_key, link in LINKS.items():
        if not link:
            continue
        for t in fetch(link, desde):
            todas.append(_normalizar(t, conta_key))
    return todas


if __name__ == "__main__" and "--test" in sys.argv:
    print(f"Agregador: {AGG}\n")
    if AGG == "arquivo":
        print(f"Pasta de extratos: {EXTRATOS_DIR.resolve()}\n")
        print(f"{'Conta':<20} {'Banco':<8} {'Agência':<8} {'Nº Conta':<16} {'Status'}")
        def _tem_arquivo(key):
            return (EXTRATOS_DIR / f"{key}.ofx").exists() or (EXTRATOS_DIR / f"{key}.xlsx").exists()

        for key, info in config.CONTAS_BANCARIAS.items():
            status = "arquivo encontrado" if _tem_arquivo(key) else "FALTA .ofx/.xlsx"
            print(f"{key:<20} {info['banco']:<8} {info['agencia']:<8} "
                  f"{info['conta']:<16} {status}")
        faltando = [k for k in config.CONTAS_BANCARIAS if not _tem_arquivo(k)]
        if faltando:
            print(f"\n⚠️  {len(faltando)} conta(s) sem arquivo .ofx/.xlsx — exporte o extrato "
                  f"do internet banking e salve em {EXTRATOS_DIR}/<CONTA>.ofx (ou .xlsx)")
            sys.exit(0)
        try:
            txs = transacoes_ultimas_24h()
            print(f"\nOK — {len(txs)} transações novas nas últimas 24h (6 contas)")
        except Exception as e:
            print(f"\nERRO ao ler extratos: {e}")
    else:
        print(f"{'Conta':<20} {'Banco':<8} {'Agência':<8} {'Nº Conta':<16} {'Status'}")
        for key, info in config.CONTAS_BANCARIAS.items():
            status = "configurada" if LINKS.get(key) else "FALTA LINK_ID"
            print(f"{key:<20} {info['banco']:<8} {info['agencia']:<8} "
                  f"{info['conta']:<16} {status}")
        faltando = [k for k, v in LINKS.items() if not v]
        if faltando:
            print(f"\n⚠️  {len(faltando)} conta(s) sem link_id — configure no .env antes de rodar em produção.")
            sys.exit(0)
        try:
            txs = transacoes_ultimas_24h()
            print(f"\nOK — {len(txs)} transações nas últimas 24h (6 contas)")
        except Exception as e:
            print(f"\nERRO de conexão: {e}")
